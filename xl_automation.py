import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def auto_excel(filename):

    wb = xl.load_workbook(filename)
    print(wb.sheetnames)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    print(cell.value)
    cell2 = sheet.cell(2, 1)
    print(cell2.value)
    maximum_cols = sheet.max_column
    maximum_rows = sheet.max_row
    print(f'Size:({maximum_rows},{maximum_cols})')

    for row in range(2, maximum_rows+1):
        cell = sheet.cell(row, 3)
        sheet.cell(row, 4).value = float(cell.value[1:]) * 0.9

    values = Reference(sheet, min_row=2, max_row=maximum_rows, min_col=4, max_col=maximum_cols)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

auto_excel('transactions.xlsx')

    # values = Reference(sheet, min_row=2, max_row=maximum_rows, min_col=1, max_col=maximum_cols)
    # chart = BarChart()
    # chart.add_data(values)
    # sheet.add_chart(chart, 'n2')



